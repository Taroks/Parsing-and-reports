import requests
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from sqlalchemy.ext.automap import automap_base
import os
import json
import pandas as pd
from bs4 import BeautifulSoup
from config import *
import re
import datetime
import locale
from openpyxl import load_workbook, Workbook

class Database:
    def __init__(self, path=None):
        self.engine = create_engine(DATABASE_ADDR)
        self.model = automap_base()
        self.model.prepare(autoload_with=self.engine)
        self.Analytic = self.model.classes.analytic
        self.session = Session(self.engine)
        self.way = path
        self.last_added_link = ""

    def add_link(self, new_link):
        try:
            data = self.Analytic(
                data = None,
                link = new_link,
                data_of_check = "2010-01-01"
            )
            self.session.add(data)
            self.session.commit()
            self.last_added_link = new_link
            return f"Ссылка {new_link} успешно добавлена"
        except Exception as e:
            return str(e)

    def create_report(self, way):
        list_of_dates = []
        list_of_links = []
        list_of_names = []
        list_of_month_1 = []
        list_of_percentages_1 = []
        list_of_month_2 = []
        list_of_percentages_2 = []
        list_of_month_3 = []
        list_of_percentages_3 = []
        list_of_month_4 = []
        list_of_percentages_4 = []
        list_of_month_5 = []
        list_of_percentages_5 = []
        list_of_month_6 = []
        list_of_percentages_6 = []
        data = self.session.query(self.Analytic).all()
        for _ in data:
            if len(_.link) > 0 and _.link is not None:
                list_of_dates.append(_.data)
                list_of_links.append(_.link)
                list_of_names.append(_.name)
                list_of_month_1.append(_.first_month)
                list_of_percentages_1.append(_.percentage_1)
                list_of_month_2.append(_.second_month)
                list_of_percentages_2.append(_.percentage_2)
                list_of_month_3.append(_.third_month)
                list_of_percentages_3.append(_.percentage_3)
                list_of_month_4.append(_.fourth_month)
                list_of_percentages_4.append(_.percentage_4)
                list_of_month_5.append(_.fifth_month)
                list_of_percentages_5.append(_.percentage_5)
                list_of_month_6.append(_.sixth_month)
                list_of_percentages_6.append(_.percentage_6)
        df = pd.DataFrame({
            "дата публикации": list_of_dates,
            "Название": list_of_names,
            "Ссылка": list_of_links,
            "1ый месяц":list_of_month_1,
            "1ый месяц проценты":list_of_percentages_1,
            "2ой месяц":list_of_month_2,
            "2ой месяц проценты":list_of_percentages_2,
            "3ий месяц":list_of_month_3,
            "3ий месяц проценты":list_of_percentages_3,
            "4ый месяц":list_of_month_4,
            "4ый месяц проценты":list_of_percentages_4,
            "5ый месяц":list_of_month_5,
            "5ый месяц проценты":list_of_percentages_5,
            "6ой месяц":list_of_month_6,
            "6ой месяц проценты":list_of_percentages_6
        })
        df.to_excel(f"{way}/Отчет.xlsx")

    def quick_deleting(self, last_link):
        try:
            query = self.session.query(self.Analytic).filter_by(link=last_link)
            record = query.first()
            if record:
                self.session.delete(record)
                self.session.commit()
                return True
            else:
                return False
        except Exception as e:
            return False

    def search_repeating(self, record):
        try:
            query = self.session.query(self.Analytic).filter_by(link=record).all()
            if query:
                return True
            else:
                return False
        except Exception as e:
            return False

    def show_all(self):
        list_of_links = []
        query = self.session.query(self.Analytic).all()
        for _ in query:
            if _.link != "":
                list_of_links.append(_.link)
        if len(list_of_links) == 0:
            list_of_links.append("список ссылок пуст")
        return list_of_links

    def delete_chosen(self, link_to_delete):
        try:
            for _ in link_to_delete:
                query = self.session.query(self.Analytic).filter_by(link=_)
                record = query.first()
                if record:
                    self.session.delete(record)
                    self.session.commit()
            return True
        except Exception as e:
            return False

    def delete_old(self):
        date_format = '%d %B %Y'
        locale.setlocale(locale.LC_TIME, 'ru_RU.UTF-8')
        a = 0
        try:
            query = self.session.query(self.Analytic).all()
            for _ in query:
                if len(_.link) > 0 and _.link is not None:
                    if datetime.date.today() >= datetime.datetime.strptime(_.data, date_format).date() + datetime.timedelta(days=180):
                        self.session.delete(_)
                        self.session.commit()
                        a += 1
            if a == 0:
                return False
            return True
        except Exception as e:
            return a

class Pathing:
    def __init__(self, path):
        if os.path.exists(f"{os.path.dirname(__file__)}/ways.json"):
            if os.path.getsize(f"{os.path.dirname(__file__)}/ways.json"):
                with open(f"{os.path.dirname(__file__)}/ways.json", "r", encoding='utf8') as file:
                    self.path = json.load(file)
            else:
                self.path = {}
        else:
            with open(f"{os.path.dirname(__file__)}/ways.json", "w", encoding='utf8') as file:
                pass

    def save_path(self, path):
        if os.path.exists(f"{os.path.dirname(__file__)}/ways.json"):
            if os.path.getsize(f"{os.path.dirname(__file__)}/ways.json") != 0:
                return False
        with open(f"{os.path.dirname(__file__)}/ways.json", "a", encoding='utf8') as file:
            json.dump(path, file)
        return path


class Parsing:
    def __init__(self, progress_callback):
        self.engine = create_engine(DATABASE_ADDR)
        self.model = automap_base()
        self.model.prepare(autoload_with=self.engine)
        self.Analytic = self.model.classes.analytic
        self.session = Session(self.engine)
        self.list_of_links = []
        self.progress_callback = progress_callback
        self.progress = 0

    def parsing(self):
        data = self.session.query(self.Analytic).all()
        success_parsing =[]
        less_then_month = []
        iterations = []
        for _ in data:
            if len(_.link) > 0 and _.link is not None:
                iterations.append(_.link)
        total_iterations = len(iterations)
        current_iteration = 0
        date_format = '%Y-%m-%d'
        locale.setlocale(locale.LC_TIME, 'ru_RU.UTF-8')
        for _ in data:
            if len(_.link) > 0 and _.link is not None:
                if datetime.date.today() >= datetime.datetime.strptime(_.data_of_check, date_format).date() + datetime.timedelta(days=30):
                    try:
                        success_parsing.append(_.link)
                        r = requests.get(_.link)
                        print(r)
                        soup = BeautifulSoup(r.text, "lxml")
                        name = soup.find("h1", class_="Heading_heading__5e023aa1 Heading_sizeL__5e023aa1 Typography_typographyHeading2__d9294f0f")
                        data = soup.find("time", class_="DateTime_datetime__2d272a9c")
                        views = soup.find("span", attrs={"data-testid": "views"})
                        if re.search (r"K", views.get_text()):
                            veews = float(views.get_text()[:-1]) * 1000
                        else:
                            veews = int(views.get_text())
                        _.name = name.get_text()
                        _.data = data.get_text()
                        _.data_of_check = datetime.date.today()
                        if _.first_month is None:
                            _.first_month = veews
                            _.percentage_1 = 0
                            self.session.commit()
                            current_iteration += 1
                            progress = current_iteration / total_iterations * 100
                            self.progress_callback(progress)
                            continue
                        if _.second_month is None:
                            _.second_month = veews
                            if _.first_month == 0:
                                _.percentage_2 = (veews/_.first_month)*100
                            else:
                                _.percentage_2 = veews * 100
                            self.session.commit()
                            current_iteration += 1
                            progress = current_iteration / total_iterations * 100
                            self.progress_callback(progress)
                            continue
                        if _.third_month is None:
                            _.third_month = veews
                            if _.second_month == 0:
                                _.percentage_3 = (veews/_.second_month)*100
                            else:
                                _.percentage_3 = veews * 100
                            self.session.commit()
                            current_iteration += 1
                            progress = current_iteration / total_iterations * 100
                            self.progress_callback(progress)
                            continue
                        if _.fourth_month is None:
                            _.fourth_month = veews
                            if _.third_month == 0:
                                _.percentage_4 = (veews/_.third_month)*100
                            else:
                                _.percentage_4 = veews * 100
                            self.session.commit()
                            current_iteration += 1
                            progress = current_iteration / total_iterations * 100
                            self.progress_callback(progress)
                            continue
                        if _.fifth_month is None:
                            _.fifth_month = veews
                            if _.fourth_month == 0:
                                _.percentage_5 = (veews/_.fourth_month)*100
                            else:
                                _.percentage_5 = veews * 100
                            self.session.commit()
                            current_iteration += 1
                            progress = current_iteration / total_iterations * 100
                            self.progress_callback(progress)
                            continue
                        if _.sixth_month is None:
                            _.sixth_month = veews
                            if _.fifth_month == 0:
                                _.percentage_6 = (veews/_.fifth_month)*100
                            else:
                                _.percentage_6 = veews * 100
                            self.session.commit()
                            current_iteration += 1
                            progress = current_iteration / total_iterations * 100
                            self.progress_callback(progress)
                            continue
                    except Exception as e:
                        link = _.link  # Замените на свою логику получения ссылки
                        error_message = f"{link} - {str(e)}"
                        print(error_message)
                        current_iteration += 1
                        progress = current_iteration / total_iterations * 100
                        self.progress_callback(progress)
                        continue
                else:
                    link = _.link
                    less_then_month.append(_.link)
                    error_message = f"{link} - месяц не прошел"
                    print(error_message)
                    continue
        return less_then_month, success_parsing

    def up_to_date_report(self, way):
        print("aaaaaaaaaaaaaaaaaa")
        data = self.session.query(self.Analytic).all()
        success_parsing = []
        current_iteration = 0
        iterations = []
        for _ in data:
            if len(_.link) > 0 and _.link is not None:
                iterations.append(_.link)
        total_iterations = len(iterations)
        list_of_dates = []
        list_of_names = []
        list_of_links = []
        list_of_views = []
        for _ in data:
            if len(_.link) > 0 and _.link is not None:
                try:
                    success_parsing.append(_.link)
                    r = requests.get(_.link)
                    print(r)
                    soup = BeautifulSoup(r.text, "lxml")
                    name = soup.find("h1",
                                     class_="Heading_heading__5e023aa1 Heading_sizeL__5e023aa1 Typography_typographyHeading2__d9294f0f")
                    data = soup.find("time", class_="DateTime_datetime__2d272a9c")
                    views = soup.find("span", attrs={"data-testid": "views"})
                    if re.search(r"K", views.get_text()):
                        veews = float(views.get_text()[:-1]) * 1000
                    else:
                        veews = int(views.get_text())
                    list_of_dates.append(data.get_text())
                    list_of_names.append(name.get_text())
                    list_of_links.append(_.link)
                    list_of_views.append(veews)
                    current_iteration += 1
                    progress = current_iteration / total_iterations * 100
                    self.progress_callback(progress)
                except Exception as e:
                    link = _.link  # Замените на свою логику получения ссылки
                    error_message = f"{link} - {str(e)}"
                    print(error_message)
                    current_iteration += 1
                    progress = current_iteration / total_iterations * 100
                    self.progress_callback(progress)
                    continue
        df = pd.DataFrame({
            "дата публикации": list_of_dates,
            "Название": list_of_names,
            "Ссылка": list_of_links,
            f"{datetime.datetime.today().date()}":list_of_views
        })
        df.to_excel(f"{way}/Отчет up-to-date.xlsx")
        self.progress_callback(100)
        return success_parsing

    def parse_tags(self, tag, way):
        current_iteration = 0
        print(os.path.dirname(__file__))
        if os.path.exists(f"{os.path.dirname(__file__)}/известные тэги.json"):
            if os.path.getsize(f"{os.path.dirname(__file__)}/известные тэги.json") != 0:
                with open(f"{os.path.dirname(__file__)}/известные тэги.json", "r", encoding='utf8') as tags:
                    dict_of_tags = json.load(tags)
            else:
                dict_of_tags = {}
        else:
            dict_of_tags = {}
        if len(tag) == 1:
            prom = dict_of_tags.get(tag)
            tag = prom
        i = len(dict_of_tags) + 1
        if tag not in list(dict_of_tags.values()):
            dict_of_tags[i] = tag
        if os.path.exists(f"{os.path.dirname(__file__)}/известные тэги.json"):
            if os.path.getsize(f"{os.path.dirname(__file__)}/известные тэги.json") != 0:
                with open(f"{os.path.dirname(__file__)}/известные тэги.json", "w", encoding='utf8') as tags:
                    tags.write(json.dumps(dict_of_tags))
        a = ""
        tinkoff_link = f"https://secrets.tinkoff.ru/tag/{tag}"
        dict_final = {}
        r = requests.get(f"{tinkoff_link}/")
        soup = BeautifulSoup(r.text, "lxml")
        try:
            if re.search("Страница не найдена", soup.find("h1", class_="article__title second-heading").get_text()):
                tinkoff_link = f"https://secrets.tinkoff.ru/{tag}/"
                r = requests.get(f"{tinkoff_link}/")
                print(r)
                soup = BeautifulSoup(r.text, "lxml")
                print(soup.find("h1", class_="article__title second-heading"), "\n ну и что за херня")
                if soup.find("h1", class_="article__title second-heading") and re.search("Страница не найдена", soup.find("h1", class_="article__title second-heading").get_text()):
                    return "Страница не найдена"
        except Exception as e:
            print(e, type(e))
        all_pages = soup.find_all("a", class_="Pagination_listItemAnchor__bad5cc06")
        if len(all_pages) != 0:
            for _ in all_pages:
                last_page = int(_.get_text())
        else:
            last_page = 1
        print("количество страниц со статьями: ", last_page)
        list_of_names = []
        list_of_links = []
        list_of_views= []
        count_of_links = []
        progress = current_iteration / last_page * 100
        self.progress_callback(progress)
        for i in range(1, last_page+1):
            for shit in soup.find_all("main"):
                for link in shit.find_all('a'):
                    if not re.search(r"tag", str(link)) and not re.search(r"page", str(link)) and not re.search(r"privacy", str(link)) and len(str(link.get("href"))) > 1:
                        count_of_links.append(link)
            r = requests.get(f"{tinkoff_link}/page/{i}")
            if r.status_code == 200:
                print("Успешно")
            soup = BeautifulSoup(r.text, 'lxml')
            #k = soup.find_all("a", class_= "Preview_anchor__26b58efd")
            for shit in soup.find_all("main"):
                print(len(count_of_links))
                for link in shit.find_all('a'):
                    if not re.search(r"tag", str(link)) and not re.search(r"page", str(link)) and not re.search(r"privacy", str(link)) and len(str(link.get("href"))) > 1:
                        list_of_links.append(tinkoff_link + link.get('href'))
                        req = requests.get(tinkoff_link + link.get('href'))
                        soup = BeautifulSoup(req.text, 'lxml')
                        print(soup.find("div", class_="preview__views"), "/", soup.find("span", attrs={"data-testid": "views"}))
                        if soup.find("div", class_="preview__views") is not None:
                            list_of_views.append(soup.find("div", class_="preview__views").get_text())
                        elif soup.find("span", attrs={"data-testid": "views"}) is not None:
                            list_of_views.append(soup.find("span", attrs={"data-testid": "views"}).get_text())
                        else:
                            list_of_views.append("0")
                        current_iteration += 1
                        progress = current_iteration / len(count_of_links) * 100
                        self.progress_callback(progress)
                for name in shit.find_all("div", class_="Typography_typographyHeading3__5d4b0944 Preview_title__26b58efd"):
                    list_of_names.append(str(name).split(">")[1].split("<")[0])
        print("формирую отчет", len(list_of_links), len(list_of_names), len(list_of_views))
        df = pd.DataFrame({
            "Название": list_of_names,
            "Ссылка": list_of_links,
            f"Просмотры на {datetime.datetime.today().date()}": list_of_views
        })
        print("report was made")
        try:
            excel_file_path = os.path.join(way, "Отчет по разделам.xlsx")

            # # Проверяем наличие файла xlsx
            if os.path.isfile(excel_file_path):
                # Файл существует, загружаем его
                book = load_workbook(excel_file_path)

                print(book.sheetnames)
                for sheet in book:
                    print(sheet)
                    # Проверяем наличие вкладки
                    if tag in book.sheetnames:
                        # Вкладка существует, удаляем ее
                        book.remove(book[tag])
                        book.save(excel_file_path)
                        print("deleted", book.sheetnames)
            # #
            # #     work_sheet = book.create_sheet(tag)
            with pd.ExcelWriter(excel_file_path, mode="a" if os.path.exists(excel_file_path) else "w") as writer:
                df.to_excel(writer, sheet_name=tag)
            #     # Создаем новую вкладку с именем tag и данными из df
            #     writer = pd.ExcelWriter(excel_file_path, engine='openpyxl', mode="a")
            #     writer.book = book
            #     writer.sheets = {ws.title: ws for ws in book.worksheets}
            #     df.to_excel(writer, sheet_name=tag, index=False)
            #     writer.save()
            # else:
            #     # Файл не существует, создаем новый с именем way и данными из df
            #     with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
            #         df.to_excel(writer, sheet_name=tag, index=False)
            #         writer.save()
        except Exception as e:
            print("возникла ошибка:", e)